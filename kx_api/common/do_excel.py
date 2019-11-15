#__coding__:'utf-8'
#auther:ly
from openpyxl import load_workbook
from kx_api.common.my_config import MyConfig
class DoExcel:
    '''该类完成从excel中读取数据，并写回测试结果'''
    def __init__(self,file_name):
        self.file_name = file_name  #excel工作簿文件名或者路径
        self.my_config = MyConfig()


    def read_data(self,sheet_name):
        '''读取测试用例excel表格数据,数据格式 [{},{}]'''

        #打开工作簿，并返回工作簿
        wb = load_workbook(self.file_name)
        #获取表单名
        sheet = wb[sheet_name]
        test_data = []
        for i in range(2,sheet.max_row + 1):
            row_data ={}  #空字典，存储每一行的数据
            row_data['CaseId'] = sheet.cell(i,1).value
            row_data['Module'] = sheet.cell(i,2).value
            print(sheet.cell(i,3).value)
            if sheet.cell(i,3).value.find('webserverAddress') !=-1:
                row_data['url'] = sheet.cell(i,3).value.replace('webserverAddress',self.my_config.get_string('serverAddress','web_server_address'))
            elif sheet.cell(i,3).value.find('serverAddress') !=-1:
                row_data['url'] = sheet.cell(i,3).value.replace('serverAddress',self.my_config.get_string('serverAddress','server_address'))
            else:
                row_data['url'] = sheet.cell(i, 3).value
            row_data['Title'] = sheet.cell(i,4).value
            row_data['Method'] = sheet.cell(i, 5).value
            if sheet.cell(i,6).value !=None:
                if sheet.cell(i,6).value.find('#PersonPhone#') !=-1:
                    tel = self.read_tel('PersonPhone')
                    row_data['Params'] = sheet.cell(i, 6).value.replace('#PersonPhone#',tel)
                else:
                    row_data['Params'] = sheet.cell(i, 6).value
            else:
                row_data['Params'] = sheet.cell(i, 6).value
            row_data['sql'] = sheet.cell(i,7).value
            row_data['ExpectedResult'] = sheet.cell(i, 8).value
            test_data.append(row_data)
        wb.close()
        finall_data = []
        #为all就返回读取到的所有用例数据，否则就返回需要的用例数据,配置文件的section就是表单名
        CaseId = MyConfig().get_string(sheet_name, 'CaseId')
        if CaseId == 'all':
            finall_data = test_data
        else:
            for i in eval(CaseId):
                finall_data.append(test_data[i-1])

        return finall_data

    def write_data(self,row,col,value,sheet_name):
        '''写回测试结果到excel'''
        wb = load_workbook(self.file_name)
        sheet = wb[sheet_name]
        sheet.cell(row,col).value = value
        wb.save(self.file_name)
        wb.close()

    def read_tel(self,sheet_name):
        '''读取会员办卡所需的电话号码,  账单编号'''
        wb = load_workbook(self.file_name)
        sheet = wb[sheet_name]
        tel = sheet.cell(2,1).value
        return str(tel)


if __name__ == '__main__':
    p = DoExcel(r'F:\AutoTest\PosApi\kx_api\test_cases\api_case.xlsx')
    # data = p.write_data(2,8,'11111','BaseInfo')
    p.write_data(2,1,'201911130000000004','billNumber')
